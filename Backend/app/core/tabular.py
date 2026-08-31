"""CSV/XLSX reading and writing for bulk files (feature 015, ADR-115).

A pure format layer: it knows a table is headers plus rows of text, and nothing about
stations, tickets, or permissions. Every value comes back as `str` — interpreting one is the
validation layer's job, and doing it here would give the two formats a chance to disagree
about what a cell meant.

Both formats exist because both sources are real: a government roster usually arrives as
CSV, while internal maintenance happens in Excel. The writers spend their effort on the
three places Excel reinterprets a cell it was handed as text — see `TEXT_COLUMNS`, the BOM,
and `_looks_like_formula` below.
"""

import csv
import io
import zipfile
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook

CSV_EXTENSION = ".csv"
XLSX_EXTENSION = ".xlsx"
SUPPORTED_EXTENSIONS = (CSV_EXTENSION, XLSX_EXTENSION)

# Excel infers a type for every cell it opens. Left alone it reads 0912345678 as the number
# 912345678, and `contact_phone` is a ticket's match key (ADR-107) — a silently dropped
# leading zero turns a whole round-trip into "nothing matched, everything is new". Writing
# these as text-formatted cells is what prevents that on the way back out of Excel too.
TEXT_COLUMNS = frozenset(
    {"uuid", "contact_phone", "contact_name", "name", "title", "no", "floor", "room"}
)

_TEXT_CELL_FORMAT = "@"
# Excel reads a BOM-less UTF-8 CSV as the local codepage, which mangles every CJK name.
_BOM = "﻿"

# A cell whose text starts with one of these is *executed* by Excel and LibreOffice when the
# file is opened. `station.contribute` is deliberately open crowd-sourcing (ADR-111), so the
# person who writes a station name is not the person who opens the export — the writers
# neutralize the leading character rather than trusting the value (ADR-208).
_FORMULA_LEADERS = ("=", "+", "-", "@", "\t", "\r")
_CSV_FORMULA_GUARD = "'"

# An .xlsx is a zip: a 1.4 MB upload inside the 2 MB cap can expand to hundreds of MB of
# sheet XML. The archive's declared sizes are checked before anything is decompressed, so a
# file that will be rejected costs a directory read rather than a parse (ADR-209).
MAX_UNCOMPRESSED_BYTES = 32 * 1024 * 1024


class TableFormatError(ValueError):
    """The uploaded file cannot be read as a table (mapped to 400 by the endpoint)."""


@dataclass(frozen=True)
class Table:
    """A parsed file: the header row, and one dict per data row keyed by header.

    `headers` is kept separately rather than derived from the first row because a header-only
    file is a legitimate import template (ADR-119) and has no first row to derive from.
    """

    headers: tuple[str, ...]
    rows: tuple[dict[str, str], ...]


def _extension_of(filename: str) -> str:
    lowered = filename.lower()
    for extension in SUPPORTED_EXTENSIONS:
        if lowered.endswith(extension):
            return extension
    raise TableFormatError(
        f"不支援的檔案格式「{filename}」；請使用 {' 或 '.join(SUPPORTED_EXTENSIONS)}"
    )


def _clean_headers(raw_headers: list) -> tuple[str, ...]:
    """Trim, then refuse anything that cannot be mapped to a column.

    A blank or duplicated header is refused rather than repaired: both mean one column's
    values would silently vanish, and guessing which one the user meant is worse than making
    them fix the file.
    """
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    while headers and headers[-1] == "":
        headers.pop()  # trailing empty columns are an artefact of the tool that wrote the file

    if not headers:
        raise TableFormatError("檔案沒有表頭列")
    if "" in headers:
        raise TableFormatError(f"第 {headers.index('') + 1} 欄沒有欄位名稱")

    duplicates = sorted({h for h in headers if headers.count(h) > 1})
    if duplicates:
        raise TableFormatError(f"表頭有重複的欄位名稱：{'、'.join(duplicates)}")

    return tuple(headers)


def _cell_to_text(value) -> str:
    """Render one cell as text, without inventing a decimal tail.

    openpyxl hands whole numbers back as floats, so an Integer column read straight from a
    spreadsheet would arrive as "200.0" and fail its own type check.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _rows_from(
    headers: tuple[str, ...], raw_rows, max_rows: int | None = None
) -> tuple[dict[str, str], ...]:
    """Zip each raw row onto the headers, dropping rows that are entirely blank.

    Spreadsheets routinely report thousands of empty rows below the data; a short row is
    padded rather than refused, since a trailing blank column is not a structural problem.

    `max_rows` stops the read one row past the caller's limit — enough for the caller to see
    that the file is over it, without parsing the rest (ADR-209). Blank rows do not count,
    so the trailing emptiness a spreadsheet writes out never consumes the allowance.
    """
    rows = []
    for raw in raw_rows:
        values = [_cell_to_text(v) for v in raw][: len(headers)]
        values += [""] * (len(headers) - len(values))
        if not any(values):
            continue
        rows.append(dict(zip(headers, values, strict=True)))
        if max_rows is not None and len(rows) > max_rows:
            break
    return tuple(rows)


def _looks_like_formula(value: str) -> bool:
    """True when a spreadsheet would execute this text instead of displaying it.

    A negative number is the reason this is not a plain prefix test: `longitude` is routinely
    `-121.5`, and guarding that would break the column's own type check on the way back in.
    """
    if not value.startswith(_FORMULA_LEADERS):
        return False
    try:
        float(value)
    except ValueError:
        return True
    return False


def _read_csv(raw: bytes, max_rows: int | None) -> Table:
    try:
        # utf-8-sig strips our own BOM when present and is plain UTF-8 when it is not.
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise TableFormatError(
            "檔案不是 UTF-8 編碼；請在試算表軟體以「CSV UTF-8」另存後再上傳"
        ) from exc

    reader = csv.reader(io.StringIO(text))
    header_row = next(reader, None)
    if header_row is None:
        raise TableFormatError("檔案是空的")
    headers = _clean_headers(header_row)
    return Table(headers=headers, rows=_rows_from(headers, reader, max_rows))


def _check_uncompressed_size(raw: bytes) -> None:
    """Reject an archive that declares more sheet data than the parser is willing to spend.

    Read from the zip directory, so this costs no decompression — the point is to spend
    nothing on a file that is going to be refused anyway (ADR-209).
    """
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            declared = sum(entry.file_size for entry in archive.infolist())
    except zipfile.BadZipFile as exc:
        raise TableFormatError("無法讀取這個 .xlsx 檔") from exc
    if declared > MAX_UNCOMPRESSED_BYTES:
        raise TableFormatError(
            f"這個 .xlsx 解壓後超過 {MAX_UNCOMPRESSED_BYTES // 1024 // 1024} MB，請分批匯入"
        )


def _read_xlsx(raw: bytes, max_rows: int | None) -> Table:
    _check_uncompressed_size(raw)
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a family of unrelated types here
        raise TableFormatError("無法讀取這個 .xlsx 檔") from exc

    sheet = workbook.active
    raw_rows = sheet.iter_rows(values_only=True)
    header_row = next(raw_rows, None)
    if header_row is None:
        workbook.close()
        raise TableFormatError("檔案是空的")
    headers = _clean_headers(list(header_row))
    rows = _rows_from(headers, raw_rows, max_rows)
    workbook.close()
    return Table(headers=headers, rows=rows)


def read_table(raw: bytes, filename: str, *, max_rows: int | None = None) -> Table:
    """Parse an uploaded CSV or XLSX into headers plus rows of text.

    `max_rows` bounds the work: parsing stops one row past it, so the caller's own row limit
    costs a rejection rather than a full parse (ADR-209).
    """
    if not raw:
        raise TableFormatError("檔案是空的")
    if _extension_of(filename) == CSV_EXTENSION:
        return _read_csv(raw, max_rows)
    return _read_xlsx(raw, max_rows)


def _csv_safe(value) -> str:
    """Prefix a formula-looking value so the spreadsheet displays it instead of running it.

    CSV carries no cell type, so the leading apostrophe is the only place to say "text" —
    unlike `write_xlsx`, which can type the cell and leave the value untouched (ADR-208).
    """
    text = "" if value is None else str(value)
    return _CSV_FORMULA_GUARD + text if _looks_like_formula(text) else text


def write_csv(headers, rows) -> bytes:
    """Render a table as UTF-8 CSV with a BOM, so Excel opens CJK text correctly.

    The BOM does not stop Excel from re-inferring cell types when the user saves — nothing in
    a CSV can. That is what `write_xlsx` is for; this format stays available because plenty
    of external sources only speak CSV.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow([_csv_safe(header) for header in headers])
    for row in rows:
        writer.writerow([_csv_safe(row.get(header, "")) for header in headers])
    return (_BOM + buffer.getvalue()).encode()


_STRING_CELL_TYPE = "s"


def _write_row(sheet, values: list[str], text_indexes: list[int]) -> None:
    """Append one row, typing every cell as a string before Excel gets to guess.

    `number_format` cannot undo this: openpyxl decides the cell's type when the value is
    assigned, so a value starting with `=` is already a formula by the time a format is set.
    Forcing the type is what keeps the value both inert and unchanged (ADR-208).
    """
    sheet.append(values)
    for column in range(1, len(values) + 1):
        cell = sheet.cell(row=sheet.max_row, column=column)
        if cell.value not in (None, ""):
            cell.data_type = _STRING_CELL_TYPE
        if column in text_indexes:
            cell.number_format = _TEXT_CELL_FORMAT


def write_xlsx(headers, rows, *, text_columns=TEXT_COLUMNS) -> bytes:
    """Render a table as .xlsx, writing every cell as text (see `_write_row`)."""
    workbook = Workbook()
    sheet = workbook.active
    _write_row(sheet, [str(header) for header in headers], [])

    text_indexes = [i + 1 for i, header in enumerate(headers) if header in text_columns]
    for row in rows:
        values = ["" if row.get(header) is None else str(row.get(header, "")) for header in headers]
        _write_row(sheet, values, text_indexes)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
