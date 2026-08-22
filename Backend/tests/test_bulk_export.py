"""Bulk export of stations and tickets (feature 015, ADR-109/119).

Export is a second read path onto data the GraphQL layer already guards, so these tests are
mostly about it repeating the same two decisions: which rows the caller's scope reaches, and
whether each ticket's contact fields go out in the clear.
"""

import io
import os

os.environ["ENV"] = "testing"

import pytest
from fastapi import HTTPException
from geoalchemy2.shape import from_shape
from openpyxl import load_workbook
from shapely.geometry import Point, Polygon
from sqlalchemy import select

from app.core.permissions import Perm
from app.core.tabular import read_table
from app.models.auth import User
from app.models.geo import Station
from app.models.property_config import StationPropertyConfig, TaskPropertyConfig
from app.models.rbac import Permission, Role, RolePermissionAssign, UserRoleAssign
from app.models.request import Tickets
from app.models.secondary_location import SecondaryLocation
from app.models.station_property import StationProperty
from app.models.team import Team, TeamZoneAssign, WorkZone
from app.models.ticket_task import TaskProperty, TicketTask
from app.services.bulk_columns import DYNAMIC_PREFIX
from app.services.bulk_export import BulkExportError, export_stations, export_tickets

IN_ZONE = Point(121.50, 25.00)
OUT_OF_ZONE = Point(121.90, 25.40)
ZONE_POLYGON = Polygon([(121.4, 24.9), (121.6, 24.9), (121.6, 25.1), (121.4, 25.1)])


async def _grant(db, user: User, perm: Perm, scope: str, role_name: str) -> None:
    permission = (
        await db.execute(select(Permission).where(Permission.key == perm.value))
    ).scalar_one_or_none()
    if permission is None:
        permission = Permission(key=perm.value)
        db.add(permission)
        await db.flush()
    role = Role(name=role_name, kind="platform")
    db.add(role)
    await db.flush()
    db.add(RolePermissionAssign(role_uuid=role.uuid, permission_uuid=permission.uuid, scope=scope))
    db.add(UserRoleAssign(user_uuid=user.uuid, role_uuid=role.uuid))
    await db.flush()


async def _zoned_team(db) -> Team:
    team = Team(name="Hualien", type="gov")
    db.add(team)
    await db.flush()
    assigner = User(name="zone-assigner")
    db.add(assigner)
    zone = WorkZone(name="Z", geometry=from_shape(ZONE_POLYGON, srid=4326))
    db.add(zone)
    await db.flush()
    db.add(TeamZoneAssign(team_uuid=team.uuid, zone_uuid=zone.uuid, assigned_by=str(assigner.uuid)))
    await db.flush()
    return team


async def _station(db, *, name: str, point: Point, creator: User, quantity: int | None = None) -> Station:
    station = Station(
        geometry=from_shape(point, srid=4326),
        created_by=str(creator.uuid),
        type="shelter",
        name=name,
        level=0,
        visibility="public",
    )
    db.add(station)
    await db.flush()
    db.add(
        SecondaryLocation(
            geometry_uuid=str(station.uuid), location_type="address", county="花蓮縣", city="光復鄉"
        )
    )
    if quantity is not None:
        db.add(
            StationProperty(
                station_uuid=station.uuid,
                property_type="facility",
                property_name="capacity_total",
                quantity=quantity,
                status="pending",
                created_by=str(creator.uuid),
            )
        )
    await db.flush()
    return station


async def _ticket_with_task(db, *, title: str, point: Point, creator: User, phone: str = "0912345678"):
    ticket = Tickets(
        geometry=from_shape(point, srid=4326),
        created_by=str(creator.uuid),
        title=title,
        contact_name="王小明",
        contact_email="wang@example.com",
        contact_phone=phone,
        status="pending",
        priority="high",
        task_type="rescue",
        visibility="public",
    )
    db.add(ticket)
    await db.flush()
    task = TicketTask(
        ticket_uuid=ticket.uuid,
        task_type="rescue",
        task_name=f"{title} 任務",
        quantity=3,
        source="user",
        visibility="public",
        created_by=str(creator.uuid),
    )
    db.add(task)
    await db.flush()
    db.add(TaskProperty(task_uuid=task.uuid, property_name="people_count", property_value="4"))
    await db.flush()
    return ticket, task


async def _configs(db) -> None:
    db.add(StationPropertyConfig(
        station_type="shelter", property_name="capacity_total", data_type="Integer", enum_options=None
    ))
    db.add(StationPropertyConfig(
        station_type="shelter", property_name="pet_friendly", data_type="Boolean", enum_options=None
    ))
    db.add(TaskPropertyConfig(
        task_type="rescue", property_name="people_count", data_type="Integer", enum_options=None
    ))
    await db.flush()


def _parse(exported) -> object:
    return read_table(exported.content, exported.filename)


# --- stations ---


@pytest.mark.asyncio
async def test_station_export_writes_the_shared_column_layout(db):
    """Headers come from `bulk_columns`, so the file is a valid import template (ADR-119)."""
    await _configs(db)
    actor = User(name="Admin")
    db.add(actor)
    await db.flush()
    await _grant(db, actor, Perm.STATION_EXPORT, "all", "exporter")
    await _station(db, name="光復國小", point=IN_ZONE, creator=actor, quantity=120)

    table = _parse(await export_stations(db, actor=actor, station_type="shelter"))

    assert table.headers[0] == "uuid"
    assert f"{DYNAMIC_PREFIX}capacity_total" in table.headers
    assert f"{DYNAMIC_PREFIX}pet_friendly" not in table.headers  # ADR-118: not storable
    assert table.rows[0]["name"] == "光復國小"
    assert table.rows[0]["county"] == "花蓮縣"
    assert table.rows[0][f"{DYNAMIC_PREFIX}capacity_total"] == "120"


@pytest.mark.asyncio
async def test_station_export_writes_the_point_as_latitude_and_longitude(db):
    """The importer needs a coordinate pair, not GeoJSON (ADR-123)."""
    actor = User(name="Admin")
    db.add(actor)
    await db.flush()
    await _grant(db, actor, Perm.STATION_EXPORT, "all", "exporter")
    await _station(db, name="光復國小", point=IN_ZONE, creator=actor)

    row = _parse(await export_stations(db, actor=actor, station_type="shelter")).rows[0]

    assert float(row["latitude"]) == pytest.approx(25.00)
    assert float(row["longitude"]) == pytest.approx(121.50)


@pytest.mark.asyncio
async def test_zone_scoped_export_only_reaches_the_team_s_own_area(db):
    """A team admin's file must contain its responsibility area and nothing else (ADR-111)."""
    await _configs(db)
    team = await _zoned_team(db)
    actor = User(name="TeamAdmin", team_uuid=team.uuid)
    author = User(name="Someone")
    db.add_all([actor, author])
    await db.flush()
    await _grant(db, actor, Perm.STATION_EXPORT, "zone", "zoned-exporter")
    await _station(db, name="區內站", point=IN_ZONE, creator=author)
    await _station(db, name="區外站", point=OUT_OF_ZONE, creator=author)

    table = _parse(await export_stations(db, actor=actor, station_type="shelter"))

    assert [row["name"] for row in table.rows] == ["區內站"]


@pytest.mark.asyncio
async def test_export_without_the_capability_is_denied(db):
    """A field worker holds no bulk capability at all (ADR-111)."""
    actor = User(name="Member")
    db.add(actor)
    await db.flush()

    with pytest.raises(HTTPException) as exc:
        await export_stations(db, actor=actor, station_type="shelter")
    assert exc.value.status_code == 403


@pytest.mark.asyncio
async def test_an_empty_result_still_produces_the_template(db):
    """A header-only file is the intended way to start from nothing (ADR-119)."""
    await _configs(db)
    actor = User(name="Admin")
    db.add(actor)
    await db.flush()
    await _grant(db, actor, Perm.STATION_EXPORT, "all", "exporter")

    table = _parse(await export_stations(db, actor=actor, station_type="shelter"))

    assert table.rows == ()
    assert f"{DYNAMIC_PREFIX}capacity_total" in table.headers


@pytest.mark.asyncio
async def test_csv_and_xlsx_carry_the_same_content(db):
    """Format is a rendering choice; the two must never diverge in what they say."""
    await _configs(db)
    actor = User(name="Admin")
    db.add(actor)
    await db.flush()
    await _grant(db, actor, Perm.STATION_EXPORT, "all", "exporter")
    await _station(db, name="光復國小", point=IN_ZONE, creator=actor, quantity=120)

    as_csv = _parse(await export_stations(db, actor=actor, station_type="shelter", file_format="csv"))
    as_xlsx = _parse(await export_stations(db, actor=actor, station_type="shelter", file_format="xlsx"))

    assert as_csv.headers == as_xlsx.headers
    assert as_csv.rows == as_xlsx.rows


@pytest.mark.asyncio
async def test_an_unsupported_format_is_refused(db):
    """ADR-115 ships CSV and XLSX only."""
    actor = User(name="Admin")
    db.add(actor)
    await db.flush()
    await _grant(db, actor, Perm.STATION_EXPORT, "all", "exporter")

    with pytest.raises(BulkExportError, match="csv|xlsx"):
        await export_stations(db, actor=actor, station_type="shelter", file_format="json")


@pytest.mark.asyncio
async def test_xlsx_export_marks_the_phone_column_as_text(db):
    """Excel must not eat the leading zero of a match key on the way back in (ADR-115)."""
    await _configs(db)
    actor = User(name="Admin")
    db.add(actor)
    await db.flush()
    await _grant(db, actor, Perm.TICKET_EXPORT, "all", "exporter")
    await _grant(db, actor, Perm.TICKET_VIEW_PII, "all", "pii-reader")
    await _ticket_with_task(db, title="求救", point=IN_ZONE, creator=actor)

    exported = await export_tickets(db, actor=actor, task_type="rescue", file_format="xlsx")
    sheet = load_workbook(io.BytesIO(exported.content)).active
    headers = [c.value for c in sheet[1]]

    assert sheet.cell(row=2, column=headers.index("contact_phone") + 1).number_format == "@"


# --- tickets ---


@pytest.mark.asyncio
async def test_ticket_export_writes_one_row_per_task(db):
    """One row is one ticket plus one task (ADR-120)."""
    await _configs(db)
    actor = User(name="Admin")
    db.add(actor)
    await db.flush()
    await _grant(db, actor, Perm.TICKET_EXPORT, "all", "exporter")
    await _grant(db, actor, Perm.TICKET_VIEW_PII, "all", "pii-reader")
    ticket, _ = await _ticket_with_task(db, title="求救", point=IN_ZONE, creator=actor)
    db.add(
        TicketTask(
            ticket_uuid=ticket.uuid, task_type="rescue", task_name="第二個任務",
            source="user", visibility="public", created_by=str(actor.uuid),
        )
    )
    await db.flush()

    table = _parse(await export_tickets(db, actor=actor, task_type="rescue"))

    assert sorted(row["task_name"] for row in table.rows) == ["求救 任務", "第二個任務"]
    assert all(row["title"] == "求救" for row in table.rows)


@pytest.mark.asyncio
async def test_ticket_export_carries_dynamic_values_of_any_type(db):
    """`task_properties.property_value` is text, so nothing is dropped on this side."""
    await _configs(db)
    actor = User(name="Admin")
    db.add(actor)
    await db.flush()
    await _grant(db, actor, Perm.TICKET_EXPORT, "all", "exporter")
    await _grant(db, actor, Perm.TICKET_VIEW_PII, "all", "pii-reader")
    await _ticket_with_task(db, title="求救", point=IN_ZONE, creator=actor)

    row = _parse(await export_tickets(db, actor=actor, task_type="rescue")).rows[0]

    assert row[f"{DYNAMIC_PREFIX}people_count"] == "4"


@pytest.mark.asyncio
async def test_pii_is_masked_per_row_inside_one_file(db):
    """The zone tier guards the export button too, row by row (ADR-109)."""
    await _configs(db)
    team = await _zoned_team(db)
    actor = User(name="TeamAdmin", team_uuid=team.uuid)
    author = User(name="Someone")
    db.add_all([actor, author])
    await db.flush()
    await _grant(db, actor, Perm.TICKET_EXPORT, "all", "exporter")
    await _grant(db, actor, Perm.TICKET_VIEW_PII, "zone", "zoned-pii")
    await _ticket_with_task(db, title="區內", point=IN_ZONE, creator=author)
    await _ticket_with_task(db, title="區外", point=OUT_OF_ZONE, creator=author)

    rows = {r["title"]: r for r in _parse(await export_tickets(db, actor=actor, task_type="rescue")).rows}

    assert rows["區內"]["contact_phone"] == "0912345678"
    assert rows["區內"]["contact_name"] == "王小明"
    assert rows["區外"]["contact_phone"] != "0912345678"
    assert rows["區外"]["contact_name"] == "王◯◯"


@pytest.mark.asyncio
async def test_a_caller_without_view_pii_gets_every_contact_masked(db):
    """Holding export does not imply holding view_pii."""
    await _configs(db)
    actor = User(name="Admin")
    db.add(actor)
    await db.flush()
    await _grant(db, actor, Perm.TICKET_EXPORT, "all", "exporter")
    await _ticket_with_task(db, title="求救", point=IN_ZONE, creator=actor)

    row = _parse(await export_tickets(db, actor=actor, task_type="rescue")).rows[0]

    assert row["contact_name"] == "王◯◯"
    assert row["contact_phone"] != "0912345678"
