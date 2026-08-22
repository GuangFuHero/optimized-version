"""CSV/XLSX read and write for bulk files (feature 015, ADR-115).

Pure format layer: it knows a table is headers plus rows of text, and nothing about
stations, tickets or permissions. Everything comes back as `str` — interpreting a value is
the validation layer's job, and doing it here would give the two formats a chance to
disagree about what a cell meant.
"""

import io
import os

os.environ["ENV"] = "testing"

import pytest
from openpyxl import Workbook, load_workbook

from app.core.tabular import (
    TEXT_COLUMNS,
    TableFormatError,
    read_table,
    write_csv,
    write_xlsx,
)

HEADERS = ("uuid", "name", "contact_phone", "level")
ROWS = [
    {"uuid": "", "name": "光復國小", "contact_phone": "0912345678", "level": "2"},
    {"uuid": "", "name": "Zhongshan Hall", "contact_phone": "+886912345678", "level": ""},
]


def _xlsx_bytes(rows: list[list]) -> bytes:
    """Build a workbook from raw python values, the way a real spreadsheet app would."""
    wb = Workbook()
    for row in rows:
        wb.active.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- CSV ---


def test_csv_starts_with_a_bom():
    """Excel reads a BOM-less UTF-8 CSV as the local codepage and mangles every CJK name."""
    assert write_csv(HEADERS, ROWS).startswith(b"\xef\xbb\xbf")


def test_csv_round_trips_cjk_and_a_leading_zero():
    """Our own reader keeps values as text, so nothing is lost between write and read."""
    table = read_table(write_csv(HEADERS, ROWS), "stations.csv")

    assert table.headers == HEADERS
    assert table.rows[0]["name"] == "光復國小"
    assert table.rows[0]["contact_phone"] == "0912345678"


def test_csv_reader_accepts_a_file_without_a_bom():
    """Files from other tools (a government roster, a `csv` module script) have no BOM."""
    raw = "name,level\n光復國小,2\n".encode()

    table = read_table(raw, "stations.csv")

    assert table.rows == ({"name": "光復國小", "level": "2"},)


def test_a_csv_that_is_not_utf8_fails_with_a_readable_message():
    """Big5 exports are common locally; the error must name the cause, not just "decode"."""
    with pytest.raises(TableFormatError, match="UTF-8"):
        read_table("名稱\n光復國小\n".encode("big5"), "stations.csv")


# --- XLSX ---


def test_xlsx_keeps_a_leading_zero_through_a_real_spreadsheet_round_trip():
    """The whole reason XLSX exists here: contact_phone is a ticket's match key (ADR-107)."""
    table = read_table(write_xlsx(HEADERS, ROWS), "tickets.xlsx")

    assert table.rows[0]["contact_phone"] == "0912345678"


def test_xlsx_writes_the_hazardous_columns_as_text_cells():
    """A text-formatted cell is what stops Excel re-inferring the value when the user saves."""
    sheet = load_workbook(io.BytesIO(write_xlsx(HEADERS, ROWS))).active
    phone_column = HEADERS.index("contact_phone") + 1

    assert sheet.cell(row=2, column=phone_column).number_format == "@"
    assert "contact_phone" in TEXT_COLUMNS


def test_xlsx_numbers_do_not_come_back_with_a_decimal_tail():
    """Whole numbers come back from openpyxl as floats; `200.0` would fail Integer validation."""
    table = read_table(_xlsx_bytes([["level"], [200]]), "stations.xlsx")

    assert table.rows[0]["level"] == "200"


def test_xlsx_blank_cells_become_empty_strings():
    """Blank means "leave this field alone" on update (ADR-121) — it must not become "None"."""
    table = read_table(_xlsx_bytes([["name", "level"], ["光復國小", None]]), "stations.xlsx")

    assert table.rows[0]["level"] == ""


def test_xlsx_trailing_blank_rows_are_dropped():
    """Excel commonly reports thousands of empty rows below the data."""
    table = read_table(
        _xlsx_bytes([["name"], ["光復國小"], [None], [None]]), "stations.xlsx"
    )

    assert len(table.rows) == 1


# --- headers ---


def test_an_empty_template_still_reports_its_headers():
    """A header-only file is a valid template (ADR-119), so headers cannot come from a row."""
    table = read_table(write_csv(HEADERS, []), "stations.csv")

    assert table.headers == HEADERS
    assert table.rows == ()


def test_duplicate_headers_are_refused():
    """Two columns of the same name silently drop one — refuse rather than pick."""
    with pytest.raises(TableFormatError, match="name"):
        read_table(b"name,name\na,b\n", "stations.csv")


def test_a_blank_header_is_refused():
    """A nameless column cannot be mapped to anything."""
    with pytest.raises(TableFormatError):
        read_table(b"name,,level\na,b,c\n", "stations.csv")


def test_headers_are_trimmed():
    """Hand-edited files pick up stray spaces around headers."""
    table = read_table(b"  name , level \na,2\n", "stations.csv")

    assert table.headers == ("name", "level")


def test_a_file_with_no_rows_at_all_is_refused():
    """A zero-byte upload is a mistake, not an empty template."""
    with pytest.raises(TableFormatError):
        read_table(b"", "stations.csv")


# --- dispatch ---


@pytest.mark.parametrize("filename", ["stations.CSV", "stations.XlsX"])
def test_extension_matching_is_case_insensitive(filename):
    """Windows and macOS both hand back mixed-case extensions."""
    raw = write_csv(HEADERS, ROWS) if filename.lower().endswith(".csv") else write_xlsx(HEADERS, ROWS)

    assert read_table(raw, filename).headers == HEADERS


@pytest.mark.parametrize("filename", ["stations.txt", "stations.json", "stations.md", "stations"])
def test_unsupported_extensions_are_refused(filename):
    """ADR-115 ships CSV and XLSX only; .md and .json were explicitly declined."""
    with pytest.raises(TableFormatError, match="csv|xlsx"):
        read_table(b"anything", filename)
